#!/usr/bin/env python3
"""Original agent loads without secrets and keeps the uploaded XPath set."""
from __future__ import annotations

import os
import sys
import unittest
from pathlib import Path

ROOT = Path(__file__).resolve().parents[1]
if str(ROOT) not in sys.path:
    sys.path.insert(0, str(ROOT))

import imds_agent_v2


class OriginalAgentTests(unittest.TestCase):
    def test_source_has_no_hardcoded_imds_password(self):
        text = (ROOT / "imds_agent_v2.py").read_text(encoding="utf-8")
        self.assertNotIn("jowk0001", text)
        self.assertIn("load_live_credentials", text)
        self.assertIn("XP_FORWARD_ACTION1", text)
        self.assertIn("XP_FIRST_RESULT_NAME", text)
        self.assertIn("wait_for_mds_content_page", text)
        self.assertIn("XP_INGREDIENTS_EXPAND", text)
        self.assertIn("XP_CONTACT_FALLBACKS", text)
        self.assertIn("def accept_passed_mds", text)
        self.assertIn("def reject_failed_mds", text)
        self.assertIn("a:has-text('Login'):visible", text)
        self.assertIn("previous-version forward prompt", text)
        self.assertIn("def is_save_changes_prompt", text)
        self.assertIn("Save-changes prompt is showing", text)
        self.assertIn("def wait_for_forwarded_own_mds", text)
        self.assertIn("Completing contact, recipients, and propose on this new ID", text)
        self.assertNotIn("dropdown.count() == 0 or not dropdown.is_visible()", text)
        self.assertIn("retrying with all statuses", text)
        self.assertIn("not clicking Ingredients on the leftover sheet", text)
        self.assertNotIn("Looking for Yes button.", text)
        self.assertNotIn('locator("input").first.wait_for(state="visible"', text)
        self.assertIn("Action Result", text)
        self.assertIn("DEFAULT_NUM_ITERATIONS = 20", text)
        self.assertIn("def resolve_num_iterations", text)
        self.assertIn("def last_lookup_company_frame", text)
        self.assertIn("def close_company_lookup_dialogs", text)
        self.assertIn("Using newest lookupCompany iframe", text)
        self.assertIn("not clicking Search", text)
        self.assertIn("Contact already", text)
        self.assertIn("not stripping lookup dialogs", text)
        self.assertIn("def wait_for_check_results", text)
        self.assertIn("def read_check_results_text", text)
        self.assertIn("def return_to_inbox_results", text)
        self.assertIn("Check waiter missed the panel", text)
        self.assertIn("Continuing with remaining rows", text)
        self.assertIn("def wait_for_connectivity", text)
        self.assertIn("def ensure_imds_session", text)
        self.assertIn("Waiting to reconnect, then retrying this row", text)
        self.assertIn("Recipient [", text)
        self.assertIn("Contact display value:", text)
        self.assertIn("Propose Failed (Contact must be specified)", text)
        self.assertIn("Contact person selection Failed", text)
        self.assertIn("close_check_results_dialog", text)
        self.assertIn("not clicking Ingredients on a leftover or search page", text)
        self.assertIn("not clicking the disabled Propose confirm", text)
        self.assertIn("clicking No so the leftover own MDS is discarded", text)
        self.assertIn("Preferred contact", text)
        self.assertIn("def contact_option_is_usable", text)
        self.assertIn("def parse_contact_option_names", text)
        self.assertIn("allow_fallback", text)
        self.assertIn("save_changes=\"no\"", text)
        self.assertIn("save-changes leave-sheet", text)
        self.assertNotIn('page.frame_locator("iframe[src*=\'lookupCompany\']")', text)
        self.assertNotIn("Fallback: using first visible iframe.", text)
        self.assertNotIn('wait_for_selector("table:has-text(\'Message\')"', text)

    def test_load_live_credentials_requires_secrets(self):
        saved = {k: os.environ.pop(k, None) for k in ("IMDS_USERNAME", "IMDS_PASSWORD", "OTP_SECRET", "IMDS_MASTER_KEY")}
        os.environ["IMDS_SKIP_VAULT"] = "1"
        try:
            with self.assertRaises(RuntimeError) as ctx:
                imds_agent_v2.load_live_credentials()
            self.assertIn("IMDS_USERNAME", str(ctx.exception))
        finally:
            os.environ.pop("IMDS_SKIP_VAULT", None)
            for key, value in saved.items():
                if value is None:
                    os.environ.pop(key, None)
                else:
                    os.environ[key] = value


class NumIterationsTests(unittest.TestCase):
    def test_empty_and_invalid_default_to_twenty(self):
        self.assertEqual(imds_agent_v2.resolve_num_iterations(""), 20)
        self.assertEqual(imds_agent_v2.resolve_num_iterations("  "), 20)
        self.assertEqual(imds_agent_v2.resolve_num_iterations("abc"), 20)
        self.assertEqual(imds_agent_v2.resolve_num_iterations("0"), 20)
        self.assertEqual(imds_agent_v2.resolve_num_iterations("-1"), 20)

    def test_leftover_three_and_ten_become_twenty(self):
        saved_three = os.environ.pop("IMDS_ALLOW_THREE", None)
        saved_ten = os.environ.pop("IMDS_ALLOW_TEN", None)
        try:
            self.assertEqual(imds_agent_v2.resolve_num_iterations("3"), 20)
            self.assertEqual(imds_agent_v2.resolve_num_iterations("10"), 20)
            os.environ["IMDS_ALLOW_THREE"] = "1"
            self.assertEqual(imds_agent_v2.resolve_num_iterations("3"), 3)
            os.environ["IMDS_ALLOW_TEN"] = "1"
            self.assertEqual(imds_agent_v2.resolve_num_iterations("10"), 10)
        finally:
            if saved_three is None:
                os.environ.pop("IMDS_ALLOW_THREE", None)
            else:
                os.environ["IMDS_ALLOW_THREE"] = saved_three
            if saved_ten is None:
                os.environ.pop("IMDS_ALLOW_TEN", None)
            else:
                os.environ["IMDS_ALLOW_TEN"] = saved_ten

    def test_explicit_counts_are_honored(self):
        self.assertEqual(imds_agent_v2.resolve_num_iterations("20"), 20)
        self.assertEqual(imds_agent_v2.resolve_num_iterations("30"), 30)
        self.assertEqual(imds_agent_v2.resolve_num_iterations("5"), 5)


class ForwardPromptHelpers(unittest.TestCase):
    def test_detects_previous_version_forward_prompt(self):
        text = (
            "You just accepted an MDS where the previous version 1521938290 / 2 "
            "has been forwarded. Do you want to forward the new version as well?"
        )
        self.assertTrue(imds_agent_v2.is_forward_previous_version_prompt(text))
        self.assertFalse(imds_agent_v2.is_forward_previous_version_prompt("Clicked Inbox button"))
        self.assertFalse(imds_agent_v2.is_forward_previous_version_prompt("Forward menu"))
        self.assertFalse(
            imds_agent_v2.is_forward_previous_version_prompt("Do you want to save your changes?")
        )

    def test_detects_save_changes_prompt(self):
        self.assertTrue(imds_agent_v2.is_save_changes_prompt("Do you want to save your changes?"))
        self.assertTrue(imds_agent_v2.is_save_changes_prompt("MDS - MATERIAL DATA SYSTEM\nDo you want to save your changes?\nYes\nNo\nCancel"))
        self.assertFalse(imds_agent_v2.is_save_changes_prompt("Do you want to forward the new version as well?"))
        self.assertFalse(imds_agent_v2.is_save_changes_prompt("Clicked Inbox button"))

    def test_mds_id_matches_numeric_id_only(self):
        self.assertTrue(imds_agent_v2.mds_id_matches("1522070544 / 2", "1522070544"))
        self.assertTrue(imds_agent_v2.mds_id_matches("1522070544 / 2.00", "1522070544"))
        self.assertTrue(imds_agent_v2.mds_id_matches("1503991331 / 0.02", "1503991331"))
        self.assertTrue(imds_agent_v2.mds_id_matches("1521938290 / 1", "1521938290"))
        self.assertFalse(imds_agent_v2.mds_id_matches("1522107776 / 1.01", "1521938290"))
        self.assertFalse(imds_agent_v2.mds_id_matches("1522107776 / 1.01", "1430442417"))
        self.assertFalse(imds_agent_v2.mds_id_matches(None, "1522070544"))
        self.assertEqual(imds_agent_v2.mds_open_status(None, "1522070544"), "unknown")
        self.assertEqual(imds_agent_v2.mds_open_status("EXTRACTION_FAILED", "1522070544"), "unknown")
        self.assertEqual(imds_agent_v2.mds_open_status("1522070544 / 2", "1522070544"), "match")
        self.assertEqual(imds_agent_v2.mds_open_status("1522107776 / 1.01", "1522070544"), "mismatch")
        self.assertEqual(imds_agent_v2.parse_mds_id_number("1522107776 / 1.01"), "1522107776")
        self.assertNotEqual(
            imds_agent_v2.parse_mds_id_number("1522275960 / 0.01"),
            imds_agent_v2.parse_mds_id_number("1522267651 / 1"),
        )


class SummaryExportTests(unittest.TestCase):
    def test_summary_columns_drop_status_and_add_action_result(self):
        self.assertIn("Action Result", imds_agent_v2.SUMMARY_COLUMNS)
        self.assertNotIn("Status", imds_agent_v2.SUMMARY_COLUMNS)

    def test_save_check_summary_writes_action_result(self):
        import tempfile
        from openpyxl import load_workbook

        rows = [{
            "MDS ID / Version": "1522247238 / 1",
            "Check Result": "Check results - 0 Error(s) / 0 Warning(s)",
            "Parts Marking Check": "PASS",
            "Recyclate Check": "PASS",
            "Biocidal Check": "PASS",
            "Overall Result": "PASS",
            "Supplier Code": "606165",
            "Part/Item No.": "1428-1130007",
            "Action Result": "Accepted, forwarded, proposed",
        }]
        with tempfile.TemporaryDirectory() as tmp:
            dest = Path(tmp) / "check_summary.xlsx"
            written = imds_agent_v2.save_check_summary(rows, dest)
            self.assertEqual(written, dest)
            wb = load_workbook(dest)
            ws = wb.active
            headers = [c.value for c in next(ws.iter_rows(min_row=1, max_row=1))]
            self.assertEqual(headers, list(imds_agent_v2.SUMMARY_COLUMNS))
            values = [c.value for c in next(ws.iter_rows(min_row=2, max_row=2))]
            self.assertEqual(values[-1], "Accepted, forwarded, proposed")
            self.assertNotIn("No", values)


class CompanyLookupHelpers(unittest.TestCase):
    def test_empty_search_criteria_prompt(self):
        self.assertTrue(
            imds_agent_v2.is_empty_search_criteria_prompt("Please enter at least one search criteria!")
        )
        self.assertTrue(
            imds_agent_v2.is_empty_search_criteria_prompt(
                "MDS - MATERIAL DATA SYSTEM\nInformation\nPlease enter at least one search criteria!"
            )
        )
        self.assertFalse(imds_agent_v2.is_empty_search_criteria_prompt("Do you want to save your changes?"))
        self.assertFalse(imds_agent_v2.is_empty_search_criteria_prompt("Clicked Search button inside iframe."))

    def test_recipient_id_in_text(self):
        tree = "Johnson Electric Industrial Manufactory Limited [9994] not yet browsed (08/26/2026)"
        self.assertTrue(imds_agent_v2.recipient_id_in_text(tree, "9994"))
        self.assertFalse(imds_agent_v2.recipient_id_in_text(tree, "293798"))
        self.assertFalse(imds_agent_v2.recipient_id_in_text("", "9994"))
        self.assertFalse(imds_agent_v2.recipient_id_in_text(tree, ""))

    def test_contact_name_matches(self):
        self.assertTrue(imds_agent_v2.contact_name_matches("Qu, Theresa", "Qu, Theresa"))
        self.assertTrue(imds_agent_v2.contact_name_matches("Qu, Theresa", "Qu"))
        self.assertFalse(imds_agent_v2.contact_name_matches("-", "Qu, Theresa"))
        self.assertFalse(imds_agent_v2.contact_name_matches("", "Qu, Theresa"))
        self.assertFalse(imds_agent_v2.contact_name_matches("Please select", "Qu, Theresa"))
        option_list = "Please select\nQu, Other\nQu, Theresa\nWong, Kam Yuen"
        self.assertFalse(imds_agent_v2.contact_name_matches(option_list, "Qu, Theresa"))
        self.assertEqual(imds_agent_v2.contact_display_value(option_list), "Please select")
        self.assertTrue(imds_agent_v2.contact_is_blank("-"))
        self.assertTrue(imds_agent_v2.contact_is_blank(""))
        self.assertTrue(imds_agent_v2.contact_is_blank("Please select"))
        self.assertFalse(imds_agent_v2.contact_is_blank("Qu, Theresa"))
        self.assertFalse(imds_agent_v2.contact_option_is_usable(""))
        self.assertFalse(imds_agent_v2.contact_option_is_usable("-"))
        self.assertFalse(imds_agent_v2.contact_option_is_usable("Please select"))
        self.assertTrue(imds_agent_v2.contact_option_is_usable("Qu, Theresa"))
        self.assertTrue(imds_agent_v2.contact_option_is_usable("Liu, Minghui"))
        dump = (
            "option-list: [\n"
            "  '',\n"
            "  'Beenah, Tan',\n"
            "  'Joe, Qiao',\n"
            "  'Liu, Minghui',\n"
            "  'Qu, Theresa',\n"
            "]"
        )
        self.assertEqual(
            imds_agent_v2.parse_contact_option_names(dump),
            ["Beenah, Tan", "Joe, Qiao", "Liu, Minghui", "Qu, Theresa"],
        )
        self.assertEqual(imds_agent_v2.preferred_contact_name(""), "Qu, Theresa")
        self.assertEqual(imds_agent_v2.preferred_contact_name("Liu, Minghui"), "Liu, Minghui")
        self.assertEqual(imds_agent_v2._save_changes_choice("no"), "no")
        self.assertEqual(imds_agent_v2._save_changes_choice("yes"), "yes")
        self.assertEqual(imds_agent_v2._save_changes_choice(None), "yes")

    def test_check_errors_blocking_prompt_and_no_js_strip(self):
        check_text = (
            "Check results - 1 Error(s) / 2 Warning(s)\n"
            "Contact must be specified\n"
            "All existing errors need to be eliminated before any further processing may take place."
        )
        self.assertTrue(imds_agent_v2.is_check_errors_blocking_prompt(check_text))
        self.assertFalse(imds_agent_v2.is_check_errors_blocking_prompt("Do you want to save your changes?"))
        self.assertEqual(
            imds_agent_v2.propose_blocked_message(check_text),
            "Propose Failed (Contact must be specified)",
        )
        self.assertEqual(
            imds_agent_v2.propose_blocked_message(
                "All existing errors need to be eliminated before any further processing may take place."
            ),
            "Propose Failed (Check errors)",
        )
        self.assertIsNone(imds_agent_v2.propose_blocked_message("Clicked Propose button."))
        self.assertFalse(imds_agent_v2.should_js_strip_modal(lookup_iframes=0, dialog_text=check_text, yes_no=False))
        self.assertFalse(
            imds_agent_v2.should_js_strip_modal(
                lookup_iframes=0,
                dialog_text="Check results - 1 Error(s) / 2 Warning(s)",
                yes_no=False,
            )
        )
        pass_text = (
            "Check results - 0 Error(s) / 0 Warning(s)\n"
            "The MDS has passed all included checks. These checks do not cover all aspects "
            "of IMDS data requirements. Further manual review may be required."
        )
        self.assertFalse(imds_agent_v2.is_check_errors_blocking_prompt(pass_text))
        self.assertTrue(imds_agent_v2.is_passing_check_results_text(pass_text))
        self.assertTrue(imds_agent_v2.is_check_results_overlay_text(pass_text))
        self.assertFalse(
            imds_agent_v2.should_js_strip_modal(lookup_iframes=0, dialog_text=pass_text, yes_no=False)
        )
        self.assertEqual(
            imds_agent_v2.preferred_check_result_message(pass_text),
            "Check results - 0 Error(s) / 0 Warning(s)",
        )
        self.assertTrue(imds_agent_v2.is_check_clean(pass_text))
        self.assertTrue(imds_agent_v2.is_check_clean("0 Error(s), 0 Warning(s)"))
        self.assertFalse(imds_agent_v2.is_check_clean("Check failed"))

    def test_company_id_was_filled(self):
        self.assertTrue(imds_agent_v2.company_id_was_filled("9994", "9994"))
        self.assertTrue(imds_agent_v2.company_id_was_filled("293798", "293798"))
        self.assertFalse(imds_agent_v2.company_id_was_filled("", "9994"))
        self.assertFalse(imds_agent_v2.company_id_was_filled(None, "293798"))
        self.assertFalse(imds_agent_v2.company_id_was_filled("9994", "293798"))

    def test_should_js_strip_modal_never_when_lookup_iframes(self):
        self.assertFalse(
            imds_agent_v2.should_js_strip_modal(lookup_iframes=2, dialog_text="", yes_no=False)
        )
        self.assertFalse(
            imds_agent_v2.should_js_strip_modal(
                lookup_iframes=1,
                dialog_text="Please enter at least one search criteria!",
                yes_no=False,
            )
        )
        self.assertFalse(
            imds_agent_v2.should_js_strip_modal(
                lookup_iframes=0, dialog_text="Do you want to save your changes?", yes_no=True
            )
        )
        self.assertTrue(imds_agent_v2.should_js_strip_modal(lookup_iframes=0, dialog_text="", yes_no=False))

    def test_check_results_present_without_visible_message_table(self):
        self.assertTrue(imds_agent_v2.check_results_present("Check results - 0 Error(s) / 0 Warning(s)"))
        self.assertTrue(imds_agent_v2.check_results_present("The MDS has passed all included checks."))
        self.assertFalse(imds_agent_v2.check_results_present("Clicked Check item."))
        self.assertFalse(imds_agent_v2.check_results_present(""))
        self.assertEqual(imds_agent_v2.parse_ui_check_counts("Check results - 0 Error(s) / 0 Warning(s)"), (0, 0))
        self.assertTrue(imds_agent_v2.is_passing_check_results_text("Check results - 0 Error(s) / 0 Warning(s)"))


class NetworkResumeTests(unittest.TestCase):
    def test_network_wait_default_is_fifteen_minutes(self):
        self.assertEqual(imds_agent_v2.network_wait_seconds(""), 15 * 60)
        self.assertEqual(imds_agent_v2.network_wait_seconds("20"), 20 * 60)
        self.assertEqual(imds_agent_v2.network_wait_seconds("abc"), 15 * 60)

    def test_transient_network_error_not_xpath_timeout(self):
        self.assertTrue(imds_agent_v2.is_transient_network_error("net::ERR_INTERNET_DISCONNECTED"))
        self.assertTrue(imds_agent_v2.is_transient_network_error("net::ERR_NAME_NOT_RESOLVED"))
        self.assertTrue(imds_agent_v2.is_transient_network_error("Failed to load login page: Timeout"))
        self.assertFalse(imds_agent_v2.is_transient_network_error("Timeout 15000ms exceeded."))
        self.assertFalse(imds_agent_v2.is_transient_network_error("Row 6 not found"))

    def test_action_result_is_complete(self):
        self.assertTrue(imds_agent_v2.action_result_is_complete("Accepted, forwarded, proposed"))
        self.assertTrue(imds_agent_v2.action_result_is_complete("Rejected"))
        self.assertFalse(imds_agent_v2.action_result_is_complete("Pending action"))
        self.assertFalse(imds_agent_v2.action_result_is_complete("Open Failed"))
        self.assertFalse(imds_agent_v2.action_result_is_complete("Propose Failed (Contact must be specified)"))


if __name__ == "__main__":
    unittest.main()
